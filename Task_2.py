#imports
import requests #allows HTML requests
import bs4  #allows BeautifulSoup useful methods
import openpyxl #allows excel sheet modification
import sys  #allows sys.maxsize
from openpyxl.utils import get_column_letter

#fill_contrib_dict - takes a list of elements representing author names in the site and a dictionary to fill
#then loops through all the text attributes (the author name) of each element. If that author is not in the
#dictionary, add them and set contributions to 1, if they are, just increment their contributions
#ASSUMES DICTIONARY IS EMPTY
def fill_contrib_dict(papers_list, dictionary):

    for name in papers_list:
        true_name = remove_commas(remove_new_lines(name.text))
        if true_name not in dictionary:
            dictionary[true_name] = 1
        else:
            dictionary[true_name] += 1

#remove_new_lines - Takes string and returns it without any newlines
def remove_new_lines(string):
    string_to_return = ''
    for char in string:
        if char != '\n':
            string_to_return += char
    return string_to_return

#remove_commas - Takes string and returns it without any commas
def remove_commas(string):
    string_to_return = ''
    for char in string:
        if char != ',':
            string_to_return += char
    return string_to_return

#merge_dicts, takes 2 dictionaries with string keys and int values, puts them into one dictionary with
#values combined among same strings
#note: for purposes of this specific program, I need to use it for 3 dictionaries, but I decided that it would
#be simpler to do it two at a time (also makes function more general use)
def merge_dicts(total_dict, dict1, dict2):

    for key in dict1.keys():
        if key in dict2:
            #first case is if key appears in both dicts
            total_dict[key] = dict1[key] + dict2[key]
        else:
            #other case is if key appears in dict1 but not dict2
            total_dict[key] = dict1[key]

    #final case is if key appears in dict2 but not dict1
    for key in dict2.keys():
        if key not in dict1:
            total_dict[key] = dict2[key]

    return total_dict

#get_largest_value_dict - returns the largest value in the dictionary that's not the largest
#largest being here allows to get second-largest onward
def get_largest_value_dict(dictionary, largest):

    this_largest = 0
    for value in dictionary.values():
        if value > this_largest and value < largest:
            this_largest = value
    return this_largest

#try_for_keyerror - does a try/except for KeyError, returns status
def try_for_keyerror(dictionary, key):

    try:
        val = dictionary[key]
    except KeyError:
        return 1
    else:
        return 0

#Variables
contrib22 = dict()  #dictionaries for {name, papers} for each year
contrib23 = dict()
contrib24 = dict()
contrib_tot = dict()

#Get the site of each year as a Response Object
res22 = requests.get("https://openaccess.thecvf.com/CVPR2022?day=all")
res23 = requests.get("https://openaccess.thecvf.com/CVPR2023?day=all")
res24 = requests.get("https://openaccess.thecvf.com/CVPR2024?day=all")

#Create a soup for each site
soup22 = bs4.BeautifulSoup(res22.text, "html.parser")
soup23 = bs4.BeautifulSoup(res23.text, "html.parser")
soup24 = bs4.BeautifulSoup(res24.text, "html.parser")

#Names of contributors are organized within form elements that have the class authsearch
papers22 = soup22.find_all("form", class_='authsearch')
papers23 = soup23.find_all("form", class_='authsearch')
papers24 = soup24.find_all("form", class_='authsearch')

#fill the contribution dictionaries with the author name and how many papers they worked on
fill_contrib_dict(papers22, contrib22)
fill_contrib_dict(papers23, contrib23)
fill_contrib_dict(papers24, contrib24)

#Merge the dictionaries to find total contributions of each author
contrib_tot = merge_dicts(contrib_tot, contrib22, contrib23)
contrib_tot = merge_dicts(contrib_tot, contrib_tot, contrib24)

#Move through the dictionary and pull out the 3 largest contributors
largest_contrib = ['', '', '']  #List containing the names of the 3 largest contributors
largest_amounts = [0, 0, 0] #List containing the amount of contributions of the 3 largest contributors

#Get the 3 largest contributions made.
# TODO: This works, cause I checked, and there's no repeats of contributions big enough to make this an issue,
# but this code assumes that the 3 largest values won't have any repeats, so maybe we should make this code consider that
largest_amounts[0] = get_largest_value_dict(contrib_tot, sys.maxsize)
largest_amounts[1] = get_largest_value_dict(contrib_tot, largest_amounts[0])
largest_amounts[2] = get_largest_value_dict(contrib_tot, largest_amounts[1])
print(largest_amounts)

#Get the names of the three largest contributors
for name in contrib_tot:
    if contrib_tot[name] == largest_amounts[0]:
        largest_contrib[0] = name
    elif contrib_tot[name] == largest_amounts[1]:
        largest_contrib[1] = name
    elif contrib_tot[name] == largest_amounts[2]:
        largest_contrib[2] = name
print(largest_contrib)

#Open the notebook
wb = openpyxl.load_workbook('top_contributors.xlsx')
sheet = wb['Sheet1']

#Fill out the sheet row by row
#This is a pretty brute-force way to do it, but it worked
for this_row in range(1, 5 + 1):   #rows

    for this_column in range(1, 4 + 1):   #columns

        if this_row == 1 and this_column != 1:
            sheet.cell(row = this_row, column = this_column).value = largest_contrib[this_column - 2]

        elif this_column == 1 and this_row != 1:
            if this_row != 5:
                sheet.cell(row = this_row, column = this_column).value = str(2022 + (this_row - 2))
            else:
                sheet.cell(row=this_row, column=this_column).value = 'Total'

        elif this_row != 1 and this_column != 1:
            if this_row == 2 and not try_for_keyerror(contrib22, largest_contrib[this_column - 2]):
                sheet.cell(row = this_row, column = this_column).value = contrib22[largest_contrib[this_column - 2]]

            elif this_row == 3 and not try_for_keyerror(contrib23, largest_contrib[this_column - 2]):
                sheet.cell(row=this_row, column=this_column).value = contrib23[largest_contrib[this_column - 2]]

            elif this_row == 4 and not try_for_keyerror(contrib23, largest_contrib[this_column - 2]):
                sheet.cell(row=this_row, column=this_column).value = contrib24[largest_contrib[this_column - 2]]

            elif this_row == 5:
                sheet.cell(row = this_row, column = this_column).value = largest_amounts[this_column - 2]

wb.save('top_contributors.xlsx')